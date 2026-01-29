import os
from openpyxl import load_workbook

def create_folders():
    # ================== 用户配置区域 ==================
    excel_path = ''  # 单引号内添加为实际路径，./xxx表示当前目录下的文件，/xxx表示根目录下的文件
    sheet_name = 'Sheet1'               # 工作表名称
    key_column = 'A'                    # 关键列（字母或数字）
    save_root = r''         # 目标目录路径

    # 标准子文件夹结构（与你提供的树状图完全一致）
    template_folders = [
        '0、前后向合同',
        '1、客情掌握',
        '2、方案总控',
        '3、谈判应标自主',
        '4、采购自主',
        '5、项目强管理',
        '6、运维自主'
    ]
    # ================================================

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # 智能检测列位置
    col_index = ord(key_column.upper()) - 64 if isinstance(key_column, str) else key_column

    success_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):  # 从第2行开始
        main_folder = str(row[col_index-1]).strip()
        
        # 过滤非法字符（保留中文括号）
        safe_name = ''.join([c if c not in '\\/:*?"<>|' else '＃' for c in main_folder])
        main_path = os.path.join(save_root, safe_name)

        if not os.path.exists(main_path):
            try:
                # 创建主文件夹
                os.makedirs(main_path)
                
                # 创建标准子文件夹
                for sub in template_folders:
                    sub_path = os.path.join(main_path, sub)
                    os.makedirs(sub_path, exist_ok=True)  # 安全模式

                success_count += 1
                print(f"✅ 成功创建：{safe_name}（含{len(template_folders)}个标准子文件夹）")
                
            except Exception as e:
                print(f"❌ 创建失败：{safe_name} - {str(e)}")
        else:
            print(f"⏩ 已存在：{safe_name}")

    print(f"\n🎉 运行结束！成功初始化 {success_count} 个项目文件夹")
    print(f"标准子文件夹结构：\n└── " + '\n└── '.join(template_folders))

if __name__ == "__main__":
    create_folders()